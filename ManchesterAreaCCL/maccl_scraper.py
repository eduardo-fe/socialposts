#!/usr/bin/env python3
"""
MACCL Results Scraper  –  fully corrected
==========================================
Requirements:
    pip install playwright openpyxl
    playwright install chromium

Key design decisions:
  1. Only scrapes the main "FINAL RESULTS" (or "PROVISIONAL RESULTS") section
     for each race group.  Age-category breakdowns (e.g. "Senior Ladies aged
     35 to 39") and TEAM RESULTS sections are deliberately skipped – they are
     either duplicates or aggregates.

  2. Two row formats are handled by one regex:
       Junior races  – Pos  Num  [$]  Name  Club  Time
       Senior races  – Pos  Num  [$]  Name  Cat  CatPos  Club  Time
     where Cat = age-category code (L20, L35, L40…L80, V40…V80, U20)
     and CatPos = finishing position within that age category.

  3. URL folder = year the race was held (not the season-end year).
     Naming convention:
       up to 2022/23 season  ->  maccl{YY}{N}   (e.g. maccl223)
       from 2023/24 onwards  ->  maccl{YY}-{N}  (e.g. maccl23-1)
     From 2024/25 M1 onwards the site publishes PDF only (no .htm).

  4. 2020/21 season was fully cancelled (COVID-19).
     2023/24 Match 3 (December) was cancelled.

  5. The ∾ character in older pages encodes & in club names – fixed.

Run:  python maccl_scraper_final.py
Out:  maccl_full_results.xlsx
"""

import re
import time
import html as htmllib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── URL table ──────────────────────────────────────────────────────────────────
B = "https://www.race-results.co.uk/results"
M = "https://maccl.co.uk/wp-content/uploads"

RACES = [
    # ── 2016/17 ──────────────────────────────────────────────
    {"season": "2016/17", "match": 1, "date": "15 Oct 2016", "url": f"{B}/2016/maccl161.htm"},
    {"season": "2016/17", "match": 2, "date": "12 Nov 2016", "url": f"{B}/2016/maccl162.htm"},
    {"season": "2016/17", "match": 3, "date": "3 Dec 2016",  "url": f"{B}/2016/maccl163.htm"},
    {"season": "2016/17", "match": 4, "date": "14 Jan 2017", "url": f"{B}/2017/maccl164.htm"},
    {"season": "2016/17", "match": 5, "date": "11 Feb 2017", "url": f"{B}/2017/maccl165.htm"},
    # ── 2017/18 ──────────────────────────────────────────────
    {"season": "2017/18", "match": 1, "date": "14 Oct 2017", "url": f"{B}/2017/maccl171.htm"},
    {"season": "2017/18", "match": 2, "date": "11 Nov 2017", "url": f"{B}/2017/maccl172.htm"},
    {"season": "2017/18", "match": 3, "date": "2 Dec 2017",  "url": f"{B}/2017/maccl173.htm"},
    {"season": "2017/18", "match": 4, "date": "13 Jan 2018", "url": f"{B}/2018/maccl174.htm"},
    {"season": "2017/18", "match": 5, "date": "10 Feb 2018", "url": f"{B}/2018/maccl175.htm"},
    # ── 2018/19 ──────────────────────────────────────────────
    {"season": "2018/19", "match": 1, "date": "13 Oct 2018", "url": f"{B}/2018/maccl181.htm"},
    {"season": "2018/19", "match": 2, "date": "10 Nov 2018", "url": f"{B}/2018/maccl182.htm"},
    {"season": "2018/19", "match": 3, "date": "1 Dec 2018",  "url": f"{B}/2018/maccl183.htm"},
    {"season": "2018/19", "match": 4, "date": "12 Jan 2019", "url": f"{B}/2019/maccl184.htm"},
    {"season": "2018/19", "match": 5, "date": "9 Feb 2019",  "url": f"{B}/2019/maccl185.htm"},
    # ── 2019/20 ──────────────────────────────────────────────
    {"season": "2019/20", "match": 1, "date": "12 Oct 2019", "url": f"{B}/2019/maccl191.htm"},
    {"season": "2019/20", "match": 2, "date": "9 Nov 2019",  "url": f"{B}/2019/maccl192.htm"},
    {"season": "2019/20", "match": 3, "date": "7 Dec 2019",  "url": f"{B}/2019/maccl193.htm"},
    {"season": "2019/20", "match": 4, "date": "11 Jan 2020", "url": f"{B}/2020/maccl194.htm"},
    {"season": "2019/20", "match": 5, "date": "8 Feb 2020",  "url": f"{B}/2020/maccl195.htm"},
    # 2020/21: ENTIRE SEASON CANCELLED (COVID-19) – omitted
    # ── 2021/22 ──────────────────────────────────────────────
    {"season": "2021/22", "match": 1, "date": "9 Oct 2021",  "url": f"{B}/2021/maccl211.htm"},
    {"season": "2021/22", "match": 2, "date": "13 Nov 2021", "url": f"{B}/2021/maccl212.htm"},
    {"season": "2021/22", "match": 3, "date": "4 Dec 2021",  "url": f"{B}/2021/maccl213.htm"},
    {"season": "2021/22", "match": 4, "date": "15 Jan 2022", "url": f"{B}/2022/maccl214.htm"},
    {"season": "2021/22", "match": 5, "date": "5 Mar 2022",  "url": f"{B}/2022/maccl215.htm"},
    # ── 2022/23 ──────────────────────────────────────────────
    {"season": "2022/23", "match": 1, "date": "15 Oct 2022", "url": f"{B}/2022/maccl221.htm"},
    {"season": "2022/23", "match": 2, "date": "12 Nov 2022", "url": f"{B}/2022/maccl222.htm"},
    {"season": "2022/23", "match": 3, "date": "14 Jan 2023", "url": f"{B}/2023/maccl223.htm"},
    {"season": "2022/23", "match": 4, "date": "11 Feb 2023", "url": f"{B}/2023/maccl224.htm"},
    {"season": "2022/23", "match": 5, "date": "11 Mar 2023", "url": f"{B}/2023/maccl225.htm"},
    # ── 2023/24 ─ naming -> maccl{YY}-{N} ───────────────────
    # Dec 2023 match CANCELLED; file numbering skips it:
    #   Oct=maccl23-1, Nov=maccl23-2, Jan=maccl23-3, Feb=maccl23-4
    {"season": "2023/24", "match": 1, "date": "14 Oct 2023", "url": f"{B}/2023/maccl23-1.htm"},
    {"season": "2023/24", "match": 2, "date": "11 Nov 2023", "url": f"{B}/2023/maccl23-2.htm"},
    # match 3 (Dec 2023) CANCELLED – omitted
    {"season": "2023/24", "match": 4, "date": "13 Jan 2024", "url": f"{B}/2024/maccl23-3.htm"},
    {"season": "2023/24", "match": 5, "date": "10 Feb 2024", "url": f"{B}/2024/maccl23-4.htm"},
    # ── 2024/25 ─ PDF only ───────────────────────────────────
    {"season": "2024/25", "match": 1, "date": "12 Oct 2024", "url": f"{M}/2024/10/MACCL24_Match-1.pdf"},
    {"season": "2024/25", "match": 2, "date": "9 Nov 2024",  "url": f"{B}/2024/maccl24-2.pdf"},
    {"season": "2024/25", "match": 3, "date": "11 Jan 2025", "url": f"{B}/2025/maccl24-3.pdf"},
    {"season": "2024/25", "match": 4, "date": "8 Feb 2025",  "url": f"{B}/2025/maccl24-4.pdf"},
    # ── 2025/26 ─ PDF only ───────────────────────────────────
    {"season": "2025/26", "match": 1, "date": "Oct 2025",    "url": f"{B}/2025/maccl25-1.pdf"},
    {"season": "2025/26", "match": 2, "date": "22 Nov 2025", "url": f"{B}/2025/maccl25-2.pdf"},
    {"season": "2025/26", "match": 3, "date": "Dec 2025",    "url": f"{B}/2025/maccl25-3.pdf"},
    {"season": "2025/26", "match": 4, "date": "Jan 2026",    "url": f"{B}/2026/maccl25-4.pdf"},
    {"season": "2025/26", "match": 5, "date": "7 Feb 2026",  "url": f"{B}/2026/maccl25-5.pdf"},
]

# ── Section filtering ──────────────────────────────────────────────────────────
# We SKIP age-category sub-breakdowns and team result sections.
# All senior runners appear in full in the main "Senior Men/Ladies -- FINAL RESULTS"
# section, so sub-breakdowns like "Senior Ladies aged 35 to 39" are redundant.
_SKIP_RE = re.compile(
    r'aged\s+\d+'       # "Senior Ladies aged 35 to 39"
    r'|over\s+\d+'      # "Senior Ladies (Vets) over 70"
    r'|\(Vets\)'        # "(Vets)" sub-sections
    r'|TEAM\s+RESULTS',
    re.IGNORECASE,
)

def _should_keep_section(heading: str) -> bool:
    """Return True only for the main FINAL / PROVISIONAL RESULTS sections."""
    h = heading.upper()
    if 'FINAL RESULTS' not in h and 'PROVISIONAL RESULTS' not in h:
        return False
    return not _SKIP_RE.search(heading)

def _clean_category(raw: str) -> str:
    """Strip the FINAL/PROVISIONAL RESULTS suffix and normalise whitespace.
    Also strips the leading venue/date prefix that appears in the markdown-
    converted pages (e.g. 'at Tatton Park on Saturday 14th January 2023 ')."""
    c = re.sub(r'\s*--?\s*(?:FINAL|PROVISIONAL)\s+RESULTS.*', '', raw, flags=re.I)
    c = re.sub(r'\(\$=non-counter\)', '', c, flags=re.I)
    c = re.sub(r'\s+', ' ', c).strip()
    # Strip "at VENUE on DAY DDth MONTH YYYY " prefix from markdown pages
    c = re.sub(r'^at\s+.+?\s+on\s+\w+\s+\d{1,2}\w*\s+\w+\s+\d{4}\s+', '', c, flags=re.I)
    return c.strip()

# ── Row parser ─────────────────────────────────────────────────────────────────
# Junior format (no age-cat column):
#   Pos  Num  [$]  Name                    Club                    Time
#
# Senior format (age-cat column present for vets / age groups):
#   Pos  Num  [$]  Name         Cat  CatPos  Club                  Time
#
# Cat codes: L20, L35, L40, L45, L50, L55, L60, L65, L70, L75, L80
#            V40, V45, V50, V55, V60, V65, V70, V75, V80
#            U20
# If cat is absent the runner is "Senior" (open / under-35 / junior).
#
_ROW_RE = re.compile(
    r'(\d+)\s+'                 # 1  overall position
    r'(\d+)\s*'                 # 2  bib / race number
    r'(\$?)\s*'                 # 3  $ = non-counter flag
    r'(.+?)\s{2,}'              # 4  name  (non-greedy, ends at 2+ spaces)
    r'(?:([LVU]\d{2})\s+'       # 5  age-cat code  (optional)
    r'(\d+)\s+)?'               # 6  cat position   (optional)
    r'(.+?)\s{2,}'              # 7  club
    r'(\d+:\d+(?::\d+)?)\s*$'  # 8  time  (mm:ss or h:mm:ss)
)

def _parse_row(line: str) -> dict | None:
    line = line.strip().replace('∾', '&')
    if not line or line.startswith('Pos') or '---' in line:
        return None
    m = _ROW_RE.match(line)
    if not m:
        return None
    pos, num, nc, name, cat, cat_pos, club, time_val = m.groups()
    return {
        'pos':         int(pos),
        'num':         int(num),
        'non_counter': nc == '$',
        'name':        name.strip(),
        'age_cat':     cat or '',              # blank for open/junior
        'cat_pos':     int(cat_pos) if cat_pos else None,
        'club':        club.strip(),
        'time':        time_val.strip(),
    }

# ── Browser fetcher ────────────────────────────────────────────────────────────
def fetch(url: str) -> str | None:
    """Fetch with a real Chromium browser, bypassing the 403 anti-bot block."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"))
            page.goto(url, wait_until='networkidle', timeout=30000)
            content = page.content()
            browser.close()
            return content
    except Exception as e:
        print(f'    fetch error: {e}')
        return None

# ── Page parser ────────────────────────────────────────────────────────────────
def parse_page(html: str, info: dict) -> list[dict]:
    """Parse one results page; returns list of individual result rows."""
    html = html.replace('∾', '&').replace('&amp;', '&')

    # Venue + date from the page heading
    m = re.search(
        r'at\s+(.+?)\s+on\s+\w+\s+(\d{1,2}(?:st|nd|rd|th)?\s+\w+\s+\d{4})',
        html)
    venue     = m.group(1).strip() if m else ''
    race_date = m.group(2).strip() if m else info['date']

    # Split on section headings (both ### markdown and <h2>/<h3> HTML)
    parts = re.split(
        r'(?:###\s*|<h[23][^>]*>)\s*'
        r'(.+?(?:FINAL|PROVISIONAL)\s+RESULTS[^<\n]*)'
        r'(?:</h[23]>)?',
        html, flags=re.I)

    results = []
    i = 1
    while i < len(parts):
        heading_raw = htmllib.unescape(re.sub(r'<[^>]+>', '', parts[i])).strip()
        body        = parts[i + 1] if i + 1 < len(parts) else ''
        i += 2

        if not _should_keep_section(heading_raw):
            continue

        category = _clean_category(heading_raw)

        # Extract the result table from <pre> or ``` fenced block
        pre = re.search(r'<pre[^>]*>(.*?)</pre>', body, re.DOTALL | re.I)
        cod = re.search(r'```(.*?)```',            body, re.DOTALL)
        if pre:
            table_text = htmllib.unescape(re.sub(r'<[^>]+>', '', pre.group(1)))
        elif cod:
            table_text = cod.group(1)
        else:
            table_text = ''

        for line in table_text.splitlines():
            row = _parse_row(line)
            if row:
                row.update({
                    'season':    info['season'],
                    'match':     info['match'],
                    'date':      info['date'],
                    'venue':     venue,
                    'race_date': race_date,
                    'category':  category,
                })
                results.append(row)

    return results

# ── Excel output ───────────────────────────────────────────────────────────────
COLUMNS = [
    ('Season',       10),
    ('Match',         7),
    ('Date',         14),
    ('Venue',        25),
    ('Race Date',    18),
    ('Category',     20),
    ('Pos',           6),
    ('Num',           7),
    ('Name',         28),
    ('Club',         35),
    ('Age Cat',       8),   # L40, V50, U20, or blank for open/junior
    ('Cat Pos',       8),   # position within age cat (blank for juniors)
    ('Time',         10),
    ('Non-Counter',  12),
]

def save_excel(all_results: list, fname: str = 'maccl_full_results.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'

    hf    = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    alt   = PatternFill('solid', fgColor='EDF2F9')
    wh    = PatternFill('solid', fgColor='FFFFFF')
    dfont = Font(name='Arial', size=9)

    # Header
    for c, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, col_name)
        cell.font      = hfont
        cell.fill      = hf
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # Data
    for r, row in enumerate(all_results, 2):
        fill = alt if r % 2 == 0 else wh
        for c, v in enumerate([
            row['season'],
            row['match'],
            row['date'],
            row['venue'],
            row['race_date'],
            row['category'],
            row['pos'],
            row['num'],
            row['name'],
            row['club'],
            row['age_cat'],
            row['cat_pos'],      # None -> blank cell
            row['time'],
            'Yes' if row['non_counter'] else '',
        ], 1):
            cell = ws.cell(r, c, v)
            cell.font = dfont
            cell.fill = fill

    wb.save(fname)
    print(f'Saved {len(all_results):,} rows  ->  {fname}')


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f'MACCL Scraper  |  {len(RACES)} races\n')
    all_results, failed = [], []

    for info in RACES:
        label = f"[{info['season']} M{info['match']}]"
        print(f'{label} {info["url"]}')
        html = fetch(info['url'])
        if not html:
            print('  -> FAILED')
            failed.append(info)
        else:
            rows = parse_page(html, info)
            if rows:
                all_results.extend(rows)
                print(f'  -> {len(rows)} rows')
            else:
                print('  -> fetched OK but 0 rows parsed')
                failed.append(info)
        time.sleep(1.5)

    save_excel(all_results)

    if failed:
        print(f'\n{len(failed)} failed:')
        for f in failed:
            print(f"  {f['season']} M{f['match']}: {f['url']}")


if __name__ == '__main__':
    main()